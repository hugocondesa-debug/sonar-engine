"""AAII weekly investor-sentiment survey connector (live xlsx).

Per spec ``docs/specs/indices/financial/F4-positioning.md`` §2: weekly
bull / bear / neutral percentages from the American Association of
Individual Investors. Thursday publication.

Primary path: public xlsx at
``https://www.aaii.com/files/surveys/sentiment.xlsx``. Requires a
browser-class User-Agent (bare curl UA → 403). Layout drift is a
real risk (spec §6 warns ``AAII_UNAVAILABLE`` flag) — schema-drift
guard checks the header row Bullish/Neutral/Bearish and raises
:class:`SchemaChangedError` on any mismatch.

Closes CAL-069.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import TYPE_CHECKING, Any

import httpx
import openpyxl
import structlog
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential_jitter,
)

from sonar.connectors.cache import DEFAULT_TTL_SECONDS, ConnectorCache
from sonar.overlays.exceptions import DataUnavailableError

if TYPE_CHECKING:
    from datetime import date
    from pathlib import Path

log = structlog.get_logger()

AAII_XLSX_URL = "https://www.aaii.com/files/surveys/sentiment.xlsx"
AAII_LEGACY_XLS_URL = "https://www.aaii.com/files/surveys/sentiment.xls"
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120 Safari/537.36"
)

# Expected xlsx header row (index 3 zero-based in the SENTIMENT sheet).
EXPECTED_HEADERS: tuple[str, ...] = (
    "Date",
    "Bullish",
    "Neutral",
    "Bearish",
)

__all__ = [
    "AAII_LEGACY_XLS_URL",
    "AAII_XLSX_URL",
    "DEFAULT_USER_AGENT",
    "EXPECTED_HEADERS",
    "AaiiConnector",
    "AaiiSurveyObservation",
    "SchemaChangedError",
]


class SchemaChangedError(DataUnavailableError):
    """Raised when AAII xlsx header row no longer matches expected schema.

    Subclass of :class:`DataUnavailableError` for backward compatibility
    with F4 ``AAII_PROXY`` degraded-path flag logic — callers can catch
    the base class without special-casing the drift variant.
    """


@dataclass(frozen=True, slots=True)
class AaiiSurveyObservation:
    """AAII weekly survey row (Thursday publication)."""

    observation_date: date
    bull_pct: float
    bear_pct: float
    neutral_pct: float
    source: str = "AAII"

    @property
    def bull_minus_bear_pct(self) -> float:
        return self.bull_pct - self.bear_pct


class AaiiConnector:
    """L0 connector for AAII sentiment survey xlsx."""

    def __init__(
        self,
        cache_dir: str | Path,
        timeout: float = 30.0,
        user_agent: str = DEFAULT_USER_AGENT,
    ) -> None:
        self.cache = ConnectorCache(cache_dir)
        self.client = httpx.AsyncClient(
            timeout=timeout,
            headers={"User-Agent": user_agent, "Accept": "*/*"},
        )

    @retry(
        stop=stop_after_attempt(5),
        wait=wait_exponential_jitter(initial=1, max=60),
        retry=retry_if_exception_type((httpx.HTTPError, httpx.TimeoutException)),
    )
    async def _fetch_xlsx_bytes(self) -> bytes:
        r = await self.client.get(AAII_XLSX_URL)
        r.raise_for_status()
        return r.content

    @staticmethod
    def _parse_workbook_bytes(content: bytes) -> list[AaiiSurveyObservation]:
        """Parse AAII xlsx bytes into observations.

        Raises :class:`SchemaChangedError` when header row deviates from
        ``EXPECTED_HEADERS``. Data rows start immediately after the
        header row; dates are ``datetime.datetime``, percentages are
        decimals (0.345 = 34.5%). We skip rows where any of bull/bear/
        neutral is None (historical header padding).
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        except Exception as e:
            err = f"AAII xlsx unreadable: {e}"
            raise DataUnavailableError(err) from e

        try:
            if "SENTIMENT" not in wb.sheetnames:
                err = f"AAII xlsx missing SENTIMENT sheet; got {wb.sheetnames}"
                raise SchemaChangedError(err)
            ws = wb["SENTIMENT"]

            rows_iter = ws.iter_rows(values_only=True)
            header_row: tuple[Any, ...] | None = None
            for i, row in enumerate(rows_iter):
                if i < 10 and row[0] == "Date":
                    header_row = row
                    break
            if header_row is None:
                err = "AAII xlsx SENTIMENT sheet missing 'Date' header row"
                raise SchemaChangedError(err)

            # Assert first 4 headers match expected sequence.
            observed = tuple(str(v) if v is not None else "" for v in header_row[:4])
            if observed != EXPECTED_HEADERS:
                err = f"AAII xlsx header drift: expected {EXPECTED_HEADERS}, got {observed}"
                raise SchemaChangedError(err)

            out: list[AaiiSurveyObservation] = []
            for row in rows_iter:
                date_cell = row[0]
                if not isinstance(date_cell, datetime):
                    continue
                bull = row[1]
                neutral = row[2]
                bear = row[3]
                if not all(isinstance(v, int | float) for v in (bull, neutral, bear)):
                    continue
                out.append(
                    AaiiSurveyObservation(
                        observation_date=date_cell.date(),
                        bull_pct=float(bull) * 100.0,
                        neutral_pct=float(neutral) * 100.0,
                        bear_pct=float(bear) * 100.0,
                    )
                )
            return out
        finally:
            wb.close()

    async def fetch_aaii_sentiment(
        self,
        start_date: date,
        end_date: date,
    ) -> list[AaiiSurveyObservation]:
        """Return AAII observations within ``[start_date, end_date]``."""
        cache_key = f"aaii_xlsx:{start_date.isoformat()}:{end_date.isoformat()}"
        cached = self.cache.get(cache_key)
        if cached is not None:
            return list(cached)
        content = await self._fetch_xlsx_bytes()
        all_obs = self._parse_workbook_bytes(content)
        within = [o for o in all_obs if start_date <= o.observation_date <= end_date]
        self.cache.set(cache_key, within, ttl=DEFAULT_TTL_SECONDS)
        log.info("aaii.fetched", n=len(within))
        return within

    async def fetch_latest(
        self, observation_date: date, *, window_days: int = 21
    ) -> AaiiSurveyObservation | None:
        """AAII publishes weekly (Thursday); default window 3 weeks."""
        start = observation_date - timedelta(days=window_days)
        obs = await self.fetch_aaii_sentiment(start, observation_date)
        usable = [o for o in obs if o.observation_date <= observation_date]
        if not usable:
            return None
        usable.sort(key=lambda o: o.observation_date)
        return usable[-1]

    async def aclose(self) -> None:
        await self.client.aclose()
        self.cache.close()
