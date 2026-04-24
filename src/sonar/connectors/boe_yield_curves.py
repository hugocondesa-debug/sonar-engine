"""Bank of England yield-curves content-store connector (Sprint Q.2).

BoE publishes the GLC/BLC yield-curve archive as daily Excel
spreadsheets distributed via its public ``-/media/boe/files`` content
store — **distinct from** the IADB Akamai-guarded endpoint used by
:mod:`sonar.connectors.boe_database`. Probe 2026-04-24 (Week 11 Day 1,
`docs/backlog/probe-results/sprint-q-2-boe-ilg-spf-probe.md`) confirmed
the content-store endpoint reaches the VPS cleanly, unlike IADB.

Scope for Sprint Q.2: the **implied inflation spot curve** archive
(``glcinflationddata.zip``, sheet ``4. spot curve``) — BoE's canonical
NSS-fitted breakeven inflation (BEI) = nominal gilt spot - real (ILG)
gilt spot, zero-coupon, daily weekday cadence. Consumed by the GB
branch of ``exp_inflation_bei`` + the ``db_backed_builder`` BEI
fallback that unblocks M3 FULL for GB (CAL-EXPINF-GB-BOE-ILG-SPF).

The connector:

1. Downloads the archive zip (cached 24h per ADR-0004 §L0-caching).
2. Picks the sub-xlsx file(s) covering ``(date_start, date_end)`` —
   archives are keyed by date band (e.g. ``2025 to present.xlsx``).
3. Parses sheet ``4. spot curve`` — row 4 carries the maturity header
   (years) and row 6+ carries daily rows ``(obs_date, values_percent…)``.
4. Returns :class:`BoeBeiSpotObservation` rows — one per
   ``(observation_date, tenor)`` pair — in decimal units (e.g. 0.035
   for 3.5 %) so the downstream BEI writer stores JSON in the same
   convention as the SPF survey writer.

Out of scope for Q.2: the nominal + real sibling archives
(``glcnominalddata.zip`` / ``glcrealddata.zip``). We intentionally
ingest only the pre-fitted implied inflation series — BoE has already
done the nominal-real subtraction. The per-tenor nominal / real input
legs are tracked via the ``BEI_FITTED_IMPLIED`` flag stamped by the
writer; the raw-leg disaggregation is a Phase-2 enhancement if we
decide to dual-source against ECB HICPx swap or equivalent.
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import TYPE_CHECKING, Any, cast

import httpx
import openpyxl
import structlog
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential_jitter,
)

from sonar.connectors.cache import DEFAULT_TTL_SECONDS, ConnectorCache
from sonar.overlays.exceptions import DataUnavailableError

if TYPE_CHECKING:
    from collections.abc import Iterable
    from pathlib import Path


log = structlog.get_logger()


__all__ = [
    "BOE_INFLATION_ARCHIVE_URL",
    "BOE_NOMINAL_ARCHIVE_URL",
    "BOE_YIELD_CURVES_BASE_URL",
    "BoeBeiSpotObservation",
    "BoeNominalSpotObservation",
    "BoeYieldCurvesConnector",
]


BOE_YIELD_CURVES_BASE_URL: str = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves"
)

# Canonical implied-inflation zip (daily cadence).
BOE_INFLATION_ARCHIVE_URL: str = f"{BOE_YIELD_CURVES_BASE_URL}/glcinflationddata.zip"

# Canonical nominal-gilt zip (daily cadence) — Sprint P.2 2026-04-24.
BOE_NOMINAL_ARCHIVE_URL: str = f"{BOE_YIELD_CURVES_BASE_URL}/glcnominalddata.zip"

# Spot-curve sheet name (Excel; case-sensitive match against the
# BoE-published workbook).
_SPOT_CURVE_SHEET: str = "4. spot curve"

# Tenor (years) → 0-based column index in the **inflation** spot-curve
# sheet. Column A (0) is the observation date; column B (1) is maturity
# 2.5Y; thereafter columns step at 0.5Y up to 40Y. Verified against the
# ``2025 to present.xlsx`` inflation workbook, header row 4. We keep
# only the tenors the SONAR M3/EXPINF stack consumes (5Y, 10Y, 15Y,
# 20Y, 30Y) — the workbook carries more but they add nothing downstream.
SPOT_CURVE_TENOR_COLUMNS: dict[float, int] = {
    5.0: 6,
    10.0: 16,
    15.0: 26,
    20.0: 36,
    30.0: 56,
}

# Tenor → column for the **nominal** spot-curve sheet. The nominal
# workbook starts at 0.5Y (column B), so ``col = 2 * years``. Verified
# 2026-04-24 against ``GLC Nominal daily data_2025 to present.xlsx``
# row 4.
NOMINAL_SPOT_CURVE_TENOR_COLUMNS: dict[float, int] = {
    2.0: 4,
    3.0: 6,
    5.0: 10,
    7.0: 14,
    10.0: 20,
    15.0: 30,
    20.0: 40,
    30.0: 60,
}

_DEFAULT_TENORS: tuple[float, ...] = (5.0, 10.0, 15.0, 20.0, 30.0)

# Default nominal tenor set — 8 standard-label tenors crossing the 5-10Y
# bridge the M3 ``nominal_5y5y_bps`` derivation depends on. 8 obs ≥
# ``MIN_OBSERVATIONS = 6`` so the NSS fitter accepts the curve; 8 < 9 =
# ``MIN_OBSERVATIONS_FOR_SVENSSON`` so the fit reduces to 4-param NS
# (flag ``NSS_REDUCED``), acceptable for backfill quality uplift.
_DEFAULT_NOMINAL_TENORS: tuple[float, ...] = (2.0, 3.0, 5.0, 7.0, 10.0, 15.0, 20.0, 30.0)


# Archive-file date bands. ``end`` is ``None`` for the rolling
# "XXXX to present" file — matched against the upper bound of the
# requested range. Ordered chronologically.
_INFLATION_ARCHIVE_FILE_BANDS: tuple[tuple[str, date, date | None], ...] = (
    ("GLC Inflation daily data_1985 to 1989.xlsx", date(1985, 1, 1), date(1989, 12, 31)),
    ("GLC Inflation daily data_1990 to 1994.xlsx", date(1990, 1, 1), date(1994, 12, 31)),
    ("GLC Inflation daily data_1995 to 1999.xlsx", date(1995, 1, 1), date(1999, 12, 31)),
    ("GLC Inflation daily data_2000 to 2004.xlsx", date(2000, 1, 1), date(2004, 12, 31)),
    ("GLC Inflation daily data_2005 to 2015.xlsx", date(2005, 1, 1), date(2015, 12, 31)),
    ("GLC Inflation daily data_2016 to 2024.xlsx", date(2016, 1, 1), date(2024, 12, 31)),
    ("GLC Inflation daily data_2025 to present.xlsx", date(2025, 1, 1), None),
)

# Nominal archive bands (Sprint P.2 2026-04-24). BoE publishes a
# 1979-1984 file on the nominal side that inflation does not carry.
_NOMINAL_ARCHIVE_FILE_BANDS: tuple[tuple[str, date, date | None], ...] = (
    ("GLC Nominal daily data_1979 to 1984.xlsx", date(1979, 1, 1), date(1984, 12, 31)),
    ("GLC Nominal daily data_1985 to 1989.xlsx", date(1985, 1, 1), date(1989, 12, 31)),
    ("GLC Nominal daily data_1990 to 1994.xlsx", date(1990, 1, 1), date(1994, 12, 31)),
    ("GLC Nominal daily data_1995 to 1999.xlsx", date(1995, 1, 1), date(1999, 12, 31)),
    ("GLC Nominal daily data_2000 to 2004.xlsx", date(2000, 1, 1), date(2004, 12, 31)),
    ("GLC Nominal daily data_2005 to 2015.xlsx", date(2005, 1, 1), date(2015, 12, 31)),
    ("GLC Nominal daily data_2016 to 2024.xlsx", date(2016, 1, 1), date(2024, 12, 31)),
    ("GLC Nominal daily data_2025 to present.xlsx", date(2025, 1, 1), None),
)


@dataclass(frozen=True, slots=True)
class BoeBeiSpotObservation:
    """One daily BEI spot-curve observation — all tenors for a date.

    ``tenors`` maps ``"5Y" | "10Y" | ...`` → decimal rate (0.035 =
    3.5 %). Values are taken directly from the BoE workbook (percent)
    and divided by 100 so the writer can persist the same
    ``bei_tenors_json`` convention used by the Sprint Q.1 SPF survey
    (decimal, not bps).
    """

    country_code: str
    observation_date: date
    tenors: dict[str, float]
    source: str = "BOE_GLC_INFLATION"


@dataclass(frozen=True, slots=True)
class BoeNominalSpotObservation:
    """One daily nominal gilt spot-curve observation (Sprint P.2).

    Same decimal convention as :class:`BoeBeiSpotObservation`; carried
    as a distinct type so downstream writers can keep BEI and nominal
    persistence paths statically disjoint.
    """

    country_code: str
    observation_date: date
    tenors: dict[str, float]
    source: str = "BOE_GLC_NOMINAL"


class BoeYieldCurvesConnector:
    """L0 connector over the BoE yield-curves content-store archive.

    Not a subclass of :class:`sonar.connectors.base.BaseConnector` —
    the archive schema is multi-tenor-per-observation (a curve, not a
    scalar series), so the connector emits
    :class:`BoeBeiSpotObservation` rather than the single-tenor
    ``Observation`` model. Follows the same cache + retry pattern.
    """

    CACHE_NAMESPACE: str = "boe_yield_curves"

    def __init__(
        self,
        *,
        cache_dir: str | Path = ".cache/boe_yield_curves",
        timeout: float = 60.0,
    ) -> None:
        self.client = httpx.AsyncClient(
            timeout=timeout,
            follow_redirects=True,
            headers={
                "User-Agent": (
                    "SONAR/2.0 (+https://github.com/hugocondesa; contact: hugocondesa@pm.me)"
                ),
            },
        )
        self.cache = ConnectorCache(cache_dir)

    async def aclose(self) -> None:
        await self.client.aclose()
        self.cache.close()

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential_jitter(initial=2, max=30),
        retry=retry_if_exception_type((httpx.HTTPError, httpx.TimeoutException)),
    )
    async def _fetch_archive_bytes(self, url: str) -> bytes:
        """GET the zip archive at ``url`` (cached, 24h TTL).

        Cache key is scoped by the archive's filename so inflation and
        nominal archives do not collide.
        """
        archive_filename = url.rsplit("/", 1)[-1]
        cache_key = f"{self.CACHE_NAMESPACE}:archive:{archive_filename}"
        cached = self.cache.get(cache_key)
        if cached is not None:
            log.debug("boe_yc.cache_hit", archive=archive_filename)
            return cast("bytes", cached)
        resp = await self.client.get(url)
        resp.raise_for_status()
        body = resp.content
        if not body or not body.startswith(b"PK"):
            msg = (
                "BoE yield-curves archive did not return a zip "
                f"(first 4 bytes: {body[:4]!r}, length={len(body)})."
            )
            raise DataUnavailableError(msg)
        self.cache.set(cache_key, body, ttl=DEFAULT_TTL_SECONDS)
        log.info("boe_yc.archive_fetched", archive=archive_filename, bytes=len(body))
        return body

    async def fetch_inflation_spot_curve(
        self,
        date_start: date,
        date_end: date,
        *,
        tenors: tuple[float, ...] = _DEFAULT_TENORS,
    ) -> list[BoeBeiSpotObservation]:
        """Return daily BEI spot-curve observations for ``[start, end]``.

        Raises :class:`DataUnavailableError` when the archive cannot be
        fetched or when none of the required sub-xlsx files carry a row
        in the requested window.
        """
        if date_start > date_end:
            msg = f"date_start {date_start} is after date_end {date_end}"
            raise ValueError(msg)
        archive_bytes = await self._fetch_archive_bytes(BOE_INFLATION_ARCHIVE_URL)
        sub_files = _select_archive_files(date_start, date_end, _INFLATION_ARCHIVE_FILE_BANDS)
        if not sub_files:
            msg = (
                "No BoE inflation sub-archive covers the requested "
                f"window {date_start}..{date_end}."
            )
            raise DataUnavailableError(msg)

        collected: list[BoeBeiSpotObservation] = []
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as zf:
            for sub_name in sub_files:
                try:
                    raw = zf.read(sub_name)
                except KeyError as exc:
                    # Archive renaming by BoE — log + skip; caller will
                    # still surface an empty collected list at the end.
                    log.warning(
                        "boe_yc.archive_member_missing",
                        member=sub_name,
                        error=str(exc),
                    )
                    continue
                collected.extend(
                    _parse_spot_curve_xlsx(
                        raw,
                        sub_name=sub_name,
                        date_start=date_start,
                        date_end=date_end,
                        tenors=tenors,
                        obs_factory=_bei_obs_factory,
                        fallback_columns=SPOT_CURVE_TENOR_COLUMNS,
                    )
                )
        if not collected:
            msg = f"BoE inflation archive returned zero rows for {date_start}..{date_end}."
            raise DataUnavailableError(msg)
        collected.sort(key=lambda o: o.observation_date)
        log.info(
            "boe_yc.parsed",
            variant="inflation",
            count=len(collected),
            date_start=date_start.isoformat(),
            date_end=date_end.isoformat(),
            tenors=[f"{t:g}Y" for t in tenors],
        )
        return collected

    async def fetch_nominal_spot_curve(
        self,
        date_start: date,
        date_end: date,
        *,
        tenors: tuple[float, ...] = _DEFAULT_NOMINAL_TENORS,
    ) -> list[BoeNominalSpotObservation]:
        """Return daily nominal-gilt spot-curve observations for ``[start, end]``.

        Sprint P.2 2026-04-24. Sibling of :meth:`fetch_inflation_spot_curve`
        against the ``glcnominalddata.zip`` archive. Used by the GB
        forwards backfill (``scripts/ops/backfill_gb_forwards.py``) to
        populate ``yield_curves_forwards`` history so M3 GB sheds the
        ``INSUFFICIENT_HISTORY`` flag.
        """
        if date_start > date_end:
            msg = f"date_start {date_start} is after date_end {date_end}"
            raise ValueError(msg)
        archive_bytes = await self._fetch_archive_bytes(BOE_NOMINAL_ARCHIVE_URL)
        sub_files = _select_archive_files(date_start, date_end, _NOMINAL_ARCHIVE_FILE_BANDS)
        if not sub_files:
            msg = (
                f"No BoE nominal sub-archive covers the requested window {date_start}..{date_end}."
            )
            raise DataUnavailableError(msg)

        collected: list[BoeNominalSpotObservation] = []
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as zf:
            for sub_name in sub_files:
                try:
                    raw = zf.read(sub_name)
                except KeyError as exc:
                    log.warning(
                        "boe_yc.archive_member_missing",
                        member=sub_name,
                        error=str(exc),
                    )
                    continue
                collected.extend(
                    _parse_spot_curve_xlsx(
                        raw,
                        sub_name=sub_name,
                        date_start=date_start,
                        date_end=date_end,
                        tenors=tenors,
                        obs_factory=_nominal_obs_factory,
                        fallback_columns=NOMINAL_SPOT_CURVE_TENOR_COLUMNS,
                    )
                )
        if not collected:
            msg = f"BoE nominal archive returned zero rows for {date_start}..{date_end}."
            raise DataUnavailableError(msg)
        collected.sort(key=lambda o: o.observation_date)
        log.info(
            "boe_yc.parsed",
            variant="nominal",
            count=len(collected),
            date_start=date_start.isoformat(),
            date_end=date_end.isoformat(),
            tenors=[f"{t:g}Y" for t in tenors],
        )
        return collected


def _select_archive_files(
    date_start: date,
    date_end: date,
    bands: tuple[tuple[str, date, date | None], ...],
) -> list[str]:
    """Return the sub-xlsx names whose date band intersects ``[start, end]``."""
    out: list[str] = []
    for name, band_start, band_end in bands:
        effective_end = band_end if band_end is not None else date_end
        if band_start <= date_end and effective_end >= date_start:
            out.append(name)
    return out


def _bei_obs_factory(
    *,
    country_code: str,
    observation_date: date,
    tenors: dict[str, float],
) -> BoeBeiSpotObservation:
    return BoeBeiSpotObservation(
        country_code=country_code,
        observation_date=observation_date,
        tenors=tenors,
    )


def _nominal_obs_factory(
    *,
    country_code: str,
    observation_date: date,
    tenors: dict[str, float],
) -> BoeNominalSpotObservation:
    return BoeNominalSpotObservation(
        country_code=country_code,
        observation_date=observation_date,
        tenors=tenors,
    )


def _parse_spot_curve_xlsx(
    raw: bytes,
    *,
    sub_name: str,
    date_start: date,
    date_end: date,
    tenors: tuple[float, ...],
    obs_factory: Any,
    fallback_columns: dict[float, int],
) -> list[Any]:
    """Parse one ``GLC … daily data_*.xlsx`` → observations.

    ``obs_factory`` builds either a :class:`BoeBeiSpotObservation` or a
    :class:`BoeNominalSpotObservation`; ``fallback_columns`` is the hard-
    coded tenor→column map used when the header-row lookup misses a
    tenor (layout-drift guard — the BoE inflation + nominal workbooks
    have distinct 2.5Y-vs-0.5Y anchors).
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if _SPOT_CURVE_SHEET not in wb.sheetnames:
            log.warning(
                "boe_yc.sheet_missing",
                sub=sub_name,
                expected=_SPOT_CURVE_SHEET,
                sheets=wb.sheetnames,
            )
            return []
        ws = wb[_SPOT_CURVE_SHEET]
        header = _read_header_row(ws)
        col_map = _map_tenors_to_columns(header, tenors, fallback_columns)
        if not col_map:
            log.warning("boe_yc.no_tenor_columns", sub=sub_name, tenors=list(tenors))
            return []
        return list(
            _iter_data_rows(
                ws,
                col_map,
                date_start=date_start,
                date_end=date_end,
                obs_factory=obs_factory,
            )
        )
    finally:
        wb.close()


def _read_header_row(ws: Any) -> tuple[Any, ...]:
    """Return row 4 of the workbook — the ``years:`` header row."""
    for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
        return cast("tuple[Any, ...]", row)
    return ()


def _map_tenors_to_columns(
    header: tuple[Any, ...],
    tenors: tuple[float, ...],
    fallback_columns: dict[float, int],
) -> dict[float, int]:
    """Locate each requested tenor in the BoE ``years:`` header row.

    Falls back to ``fallback_columns`` when a tenor is not found in the
    header (layout-drift guard — BoE's workbook format has been stable
    for decades but we still protect the parse). Inflation and nominal
    archives have distinct fallback maps because their first data
    column anchors at different maturities (2.5Y vs 0.5Y).
    """
    out: dict[float, int] = {}
    for tenor in tenors:
        found: int | None = None
        for i, v in enumerate(header):
            if isinstance(v, (int, float)) and abs(float(v) - tenor) < 0.01:
                found = i
                break
        if found is None:
            found = fallback_columns.get(tenor)
        if found is None:
            log.warning("boe_yc.tenor_missing_from_header", tenor=tenor)
            continue
        out[tenor] = found
    return out


def _iter_data_rows(
    ws: Any,
    col_map: dict[float, int],
    *,
    date_start: date,
    date_end: date,
    obs_factory: Any,
) -> Iterable[Any]:
    """Yield observation rows from the spot-curve sheet.

    Rows are emitted only when **all** mapped tenors carry a numeric
    value for that date — partial rows are skipped (matches the
    ``exp_inflation_bei`` writer contract that requires a full
    ``bei_tenors_json`` payload).
    """
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row:
            continue
        obs_raw = row[0]
        obs_date = _coerce_date(obs_raw)
        if obs_date is None:
            continue
        if obs_date < date_start or obs_date > date_end:
            continue
        tenors_decimal: dict[str, float] = {}
        all_present = True
        for tenor, col_idx in col_map.items():
            if col_idx >= len(row):
                all_present = False
                break
            raw_val = row[col_idx]
            if raw_val is None:
                all_present = False
                break
            try:
                pct = float(raw_val)
            except (TypeError, ValueError):
                all_present = False
                break
            tenors_decimal[_tenor_key(tenor)] = pct / 100.0
        if not all_present or not tenors_decimal:
            continue
        yield obs_factory(
            country_code="GB",
            observation_date=obs_date,
            tenors=tenors_decimal,
        )


def _coerce_date(value: object) -> date | None:
    """Coerce a spot-curve row's column-A cell to :class:`date`.

    BoE's workbook stores observation dates as Excel datetimes;
    openpyxl surfaces them as :class:`datetime.datetime`. Guard
    against string variants (rare — legacy archives) and ``None``
    (empty trailing rows in the 2025-to-present sheet).
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%d %b %Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).replace(tzinfo=UTC).date()
            except ValueError:
                continue
    return None


def _tenor_key(tenor: float) -> str:
    """Format the tenor as ``"5Y" | "10Y" | "30Y"`` (integer-year only).

    All BoE spot-curve tenors we consume land on whole years; we keep
    the key format plain ("5Y") to match the Sprint Q.1 SPF
    ``interpolated_tenors_json`` convention (``"5Y"``, ``"10Y"``, etc.)
    so the downstream ``_bei_tenors_bps`` parser behaves like the
    survey counterpart.
    """
    as_int = round(tenor)
    return f"{as_int}Y"
