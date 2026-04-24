"""Bank of Japan (BoJ) Tankan Inflation-Outlook L0 connector.

Sibling of :mod:`sonar.connectors.boj` (which wraps the browser-gated
TSD). The Tankan "Summary of Inflation Outlook of Enterprises" lands
in a separate publication stream exposed as public ZIP archives:
:data:`BOJ_TANKAN_GAIYO_URL_TEMPLATE` → a ZIP containing
``GA_E1.xlsx``; sheet ``TABLE7`` holds the inflation-outlook table
(Sprint Q.3 empirical probe, 2026-04-24 —
``docs/backlog/probe-results/sprint-q-3-jp-ca-survey-probe.md`` §1).

Scope at Sprint Q.3: consumes the **All Enterprises / All industries /
General Prices / Current projection** row — canonical economy-wide
inflation expectation at 1Y / 3Y / 5Y horizons, one observation per
quarterly release. Data integrated into the TANKAN Summary from the
March 2020 survey onward; pre-2020 releases live in standalone bukka
PDFs (scope-locked out — ``CAL-EXPINF-JP-SCRAPE-PRE2020`` Week 12+).

Returns :class:`TankanInflationOutlook` per release — the connector
is publication-date-centric (one ``fetch_release`` per release year +
quarter) rather than observation-date-centric (contrast
:class:`~sonar.connectors.boc.BoCConnector` which dates per-day by
construction).
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from datetime import date as date_type
from typing import TYPE_CHECKING, Final

import httpx
import structlog
from openpyxl import load_workbook
from tenacity import (
    RetryError,
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential_jitter,
)

from sonar.connectors.cache import DEFAULT_TTL_SECONDS, ConnectorCache
from sonar.overlays.exceptions import DataUnavailableError

if TYPE_CHECKING:
    from pathlib import Path

log = structlog.get_logger()


__all__ = [
    "BOJ_TANKAN_GAIYO_URL_TEMPLATE",
    "BoJTankanConnector",
    "TankanInflationOutlook",
]


# URL template for the English-language Tankan Summary ZIP, per
# §1.3 of the Sprint Q.3 probe. Placeholders: ``bucket`` (5-year
# archive directory — e.g. 2021 for releases 2021-2025, current year
# for the most recent release), ``yy`` (2-digit publication year),
# ``mm`` (quarter-end month: 03 / 06 / 09 / 12). BoJ co-locates 5
# consecutive years under a single numbered directory; the most
# recent release year has its own directory. Callers probe both
# candidates via :func:`_bucket_candidates`.
BOJ_TANKAN_GAIYO_URL_TEMPLATE: Final[str] = (
    "https://www.boj.or.jp/en/statistics/tk/gaiyo/{bucket}/tka{yy}{mm}.zip"
)

# First publication year where the gaiyo ZIP (not PDF) format is
# available. Pre-2021 releases only land as PDF in the earlier 5-year
# buckets — handled by ``CAL-EXPINF-JP-SCRAPE-PRE2020`` Week 12+.
_FIRST_ZIP_YEAR: Final[int] = 2021


# Tankan Summary XLSX inside each ZIP — stable filename since
# integration (March 2020 release onward).
_TANKAN_XLSX_NAME: Final[str] = "GA_E1.xlsx"
_TANKAN_SHEET: Final[str] = "TABLE7"

# Column layout inside TABLE7 (0-indexed, per openpyxl tuple output).
# Verified empirically on tka2603.zip 2026-Q1 release.
_COL_ROW_LABEL: Final[int] = 0  # "Large Enterprises" | "Small Enterprises" | "All Enterprises"
_COL_INDUSTRY: Final[int] = 1  # "Manu-facturing" | "Nonmanu-facturing" | "All industries"
_COL_HORIZON: Final[int] = 2  # "1 year ahead" | "3 years ahead" | "5 years ahead"
_COL_PROJECTION: Final[int] = 3  # "Previous projection" | "Current projection"
_COL_GENERAL_PRICES_VALUE: Final[int] = 6  # current projection value (% per annum)

_ALL_ENTERPRISES: Final[str] = "All Enterprises"
_ALL_INDUSTRIES: Final[str] = "All industries"
_CURRENT_PROJECTION: Final[str] = "Current projection"
_HORIZON_LABELS: Final[dict[str, str]] = {
    "1 year ahead": "1Y",
    "3 years ahead": "3Y",
    "5 years ahead": "5Y",
}


@dataclass(frozen=True, slots=True)
class TankanInflationOutlook:
    """Parsed Tankan inflation outlook for a single release.

    Horizons are keyed by canonical tenor labels (``"1Y"`` / ``"3Y"``
    / ``"5Y"``) and carry the **% per annum** general-prices value
    from the "All Enterprises / All industries / General Prices /
    Current projection" row of TABLE7. ``reference_date`` is the
    first day of the quarter-end month (e.g. ``2026-03-01`` for the
    March 2026 release).
    """

    reference_date: date_type
    release_year: int
    release_quarter_end_month: int
    horizons_pct: dict[str, float]


class BoJTankanConnector:
    """L0 connector for the BoJ Tankan Summary inflation-outlook ZIP."""

    CACHE_NAMESPACE: Final[str] = "boj_tankan"
    CONNECTOR_ID: Final[str] = "boj_tankan"

    def __init__(
        self,
        cache_dir: str | Path,
        timeout: float = 30.0,
    ) -> None:
        self.cache = ConnectorCache(cache_dir)
        self.client = httpx.AsyncClient(
            timeout=timeout,
            follow_redirects=True,
            headers={
                "User-Agent": "SONAR/2.0 (monetary-cascade; contact hugocondesa@pm.me)",
                "Accept": "application/zip, */*",
            },
        )

    async def _fetch_zip(self, release_year: int, release_month: int) -> bytes:
        """Fetch the release ZIP, probing each bucket candidate in turn.

        The first candidate to return HTTP 200 wins; persistent 404s
        propagate as :class:`httpx.HTTPStatusError` out of the retry
        wrapper so the caller can tag the release unavailable.
        """
        last_exc: httpx.HTTPError | None = None
        for bucket in _bucket_candidates(release_year):
            url = BOJ_TANKAN_GAIYO_URL_TEMPLATE.format(
                bucket=bucket,
                yy=f"{release_year % 100:02d}",
                mm=f"{release_month:02d}",
            )
            try:
                return await self._fetch_bucket(url)
            except httpx.HTTPStatusError as exc:
                if exc.response.status_code != 404:
                    raise
                last_exc = exc
                continue
        if last_exc is not None:
            raise last_exc
        msg = f"no bucket candidates for release {release_year}-{release_month:02d}"
        raise httpx.HTTPError(msg)

    @retry(
        stop=stop_after_attempt(5),
        wait=wait_exponential_jitter(initial=1, max=30),
        retry=retry_if_exception(
            lambda exc: (
                isinstance(exc, (httpx.TransportError, httpx.TimeoutException))
                or (isinstance(exc, httpx.HTTPStatusError) and exc.response.status_code >= 500)
            )
        ),
    )
    async def _fetch_bucket(self, url: str) -> bytes:
        """Inner retryable GET — retries on 5xx + transport errors only.

        4xx (notably 404 for a non-existent bucket candidate) surface
        immediately so :meth:`_fetch_zip` can fall back to the next
        bucket without burning five retry attempts per miss.
        """
        r = await self.client.get(url)
        r.raise_for_status()
        return r.content

    async def fetch_release(
        self,
        release_year: int,
        release_quarter_end_month: int,
    ) -> TankanInflationOutlook:
        """Fetch + parse one Tankan release.

        ``release_year`` is the publication year (4-digit);
        ``release_quarter_end_month`` is one of 3, 6, 9, 12. Observations
        are the "All Enterprises / All industries / General Prices /
        Current projection" row of TABLE7.

        Raises :class:`DataUnavailableError` on HTTP failure, ZIP
        decode failure, missing XLSX, or parse failure.
        """
        if release_quarter_end_month not in (3, 6, 9, 12):
            msg = f"Tankan release month must be 3/6/9/12, got {release_quarter_end_month!r}"
            raise ValueError(msg)

        cache_key = f"{self.CACHE_NAMESPACE}:release:{release_year}:{release_quarter_end_month:02d}"
        cached = self.cache.get(cache_key)
        if isinstance(cached, TankanInflationOutlook):
            log.debug("boj_tankan.cache_hit", year=release_year, month=release_quarter_end_month)
            return cached

        try:
            blob = await self._fetch_zip(release_year, release_quarter_end_month)
        except (httpx.HTTPError, RetryError) as exc:
            msg = (
                f"BoJ Tankan release {release_year}-Q{release_quarter_end_month // 3}: "
                f"fetch failed — {exc}"
            )
            raise DataUnavailableError(msg) from exc

        try:
            horizons_pct = _parse_tankan_xlsx_blob(blob)
        except (KeyError, ValueError, zipfile.BadZipFile) as exc:
            msg = (
                f"BoJ Tankan release {release_year}-Q{release_quarter_end_month // 3}: "
                f"parse failed — {exc}"
            )
            raise DataUnavailableError(msg) from exc

        if not horizons_pct:
            msg = (
                f"BoJ Tankan release {release_year}-Q{release_quarter_end_month // 3}: "
                f"TABLE7 produced no horizon values"
            )
            raise DataUnavailableError(msg)

        outlook = TankanInflationOutlook(
            reference_date=date_type(release_year, release_quarter_end_month, 1),
            release_year=release_year,
            release_quarter_end_month=release_quarter_end_month,
            horizons_pct=horizons_pct,
        )
        self.cache.set(cache_key, outlook, ttl=DEFAULT_TTL_SECONDS)
        log.info(
            "boj_tankan.fetched",
            year=release_year,
            month=release_quarter_end_month,
            horizons=sorted(horizons_pct),
        )
        return outlook

    async def aclose(self) -> None:
        await self.client.aclose()
        self.cache.close()


def _parse_tankan_xlsx_blob(blob: bytes) -> dict[str, float]:
    """Extract the All-Enterprises General-Prices current-projection values.

    Scans :data:`_TANKAN_SHEET` for the "All Enterprises" block (row
    where col[0] == ``All Enterprises`` and col[1] == ``All industries``)
    then walks forward collecting ``(horizon, current-projection)``
    pairs. Parser halts at the next top-level block (col[0] non-null and
    different) or end-of-data.

    Returns a dict keyed by canonical tenor labels (``"1Y"``, ``"3Y"``,
    ``"5Y"``). Empty return → caller raises DataUnavailableError.
    """
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        if _TANKAN_XLSX_NAME not in zf.namelist():
            available = ", ".join(zf.namelist())
            msg = f"Tankan ZIP missing {_TANKAN_XLSX_NAME!r}; has: {available}"
            raise KeyError(msg)
        with zf.open(_TANKAN_XLSX_NAME) as fh:
            xlsx_bytes = fh.read()

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        if _TANKAN_SHEET not in wb.sheetnames:
            msg = f"Tankan XLSX missing sheet {_TANKAN_SHEET!r}; has: {wb.sheetnames}"
            raise KeyError(msg)
        ws = wb[_TANKAN_SHEET]

        horizons_pct: dict[str, float] = {}
        inside_all_enterprises = False
        pending_horizon: str | None = None

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            row_label = row[_COL_ROW_LABEL] if len(row) > _COL_ROW_LABEL else None
            industry = row[_COL_INDUSTRY] if len(row) > _COL_INDUSTRY else None
            horizon_raw = row[_COL_HORIZON] if len(row) > _COL_HORIZON else None
            projection = row[_COL_PROJECTION] if len(row) > _COL_PROJECTION else None
            value_raw = (
                row[_COL_GENERAL_PRICES_VALUE] if len(row) > _COL_GENERAL_PRICES_VALUE else None
            )

            if isinstance(row_label, str) and row_label.strip() == _ALL_ENTERPRISES:
                if isinstance(industry, str) and industry.strip() == _ALL_INDUSTRIES:
                    inside_all_enterprises = True
                    pending_horizon = _match_horizon(horizon_raw)
                    if (
                        pending_horizon is not None
                        and isinstance(projection, str)
                        and projection.strip() == _CURRENT_PROJECTION
                    ):
                        _try_record(horizons_pct, pending_horizon, value_raw)
                continue

            if inside_all_enterprises and isinstance(row_label, str) and row_label.strip():
                # Next top-level label — exit block.
                break

            if not inside_all_enterprises:
                continue

            horizon_match = _match_horizon(horizon_raw)
            if horizon_match is not None:
                pending_horizon = horizon_match

            if (
                pending_horizon is not None
                and isinstance(projection, str)
                and projection.strip() == _CURRENT_PROJECTION
            ):
                _try_record(horizons_pct, pending_horizon, value_raw)
    finally:
        wb.close()

    return horizons_pct


def _bucket_candidates(release_year: int) -> list[int]:
    """Candidate gaiyo/{bucket} directory numbers, in probe order.

    BoJ co-locates 5 consecutive years under a single numbered
    directory (``gaiyo/2016/`` for 2016-2020, ``gaiyo/2021/`` for
    2021-2025, …) while the current release year has its own
    directory. We probe the year-specific directory first (handles
    the always-current bucket edge case) and fall back to the 5-year
    floor (handles historical releases). Years earlier than
    :data:`_FIRST_ZIP_YEAR` return an empty list — pre-2021 data is
    PDF-only.
    """
    if release_year < _FIRST_ZIP_YEAR:
        return []
    # BoJ buckets start at years ending in 1 or 6 (… 2011, 2016, 2021,
    # 2026 …). Formula: ((y - 1) // 5) * 5 + 1 rounds down to the
    # nearest such year.
    five_year_bucket = ((release_year - 1) // 5) * 5 + 1
    candidates: list[int] = [release_year]
    if five_year_bucket != release_year:
        candidates.append(five_year_bucket)
    return candidates


def _match_horizon(cell: object) -> str | None:
    if not isinstance(cell, str):
        return None
    return _HORIZON_LABELS.get(cell.strip())


def _try_record(target: dict[str, float], horizon: str, value_raw: object) -> None:
    if horizon in target:
        return
    try:
        target[horizon] = float(value_raw)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return
