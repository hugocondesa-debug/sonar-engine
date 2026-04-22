"""Damodaran Historical + Monthly Implied ERP connector — xval reference
+ live mature-market input.

Two endpoints:

- **Annual histimpl**
  ``https://pages.stern.nyu.edu/~adamodar/pc/datasets/histimpl.xlsx``
- **Monthly implied ERP archive**
  ``https://pages.stern.nyu.edu/~adamodar/pc/implprem/ERP{MonAbbr}{YY}.xlsx``
  e.g. ``ERPFeb26.xlsx``. Each monthly workbook carries **all** prior
  monthly rows back to 2008-09 under the ``Implied ERP (Monthly from
  9-08)`` sheet, so one recent file is sufficient for historical
  lookups. Files are published with a ~2-month lag — the fetcher
  probes backward from the target month until a 200 is found.

Content structure (annual file, 2024 snapshot): workbook with sheet
``Historical Impl Premiums`` containing annual rows 1960-present.
Columns include ``Year``, ``Earnings Yield``, ``Dividend Yield``,
``S&P 500``, ``Earnings*``, ``Implied Premium (FCFE with sustainable
Payout)``, etc.

Monthly file columns (Feb 2026 snapshot): ``Start of month``,
``S&P 500``, ``T.Bond Rate``, ``$ Riskfree Rate``, ``Ten-year average
CF``, ``CF (Trailing 12 month)``, ``Normalized CF``, ``Expected growth
rate``, ``ERP (T12 m with sustainable payout)``, ``ERP (T12m)``.

Spec §4 step 8 mentions "date.month" for xval — the annual file
serves the **year-matched xval** for ERP US canonical fits, while
the monthly file serves the **live mature-market ERP input** for
:mod:`sonar.pipelines.daily_cost_of_capital` (replacing the static
``DAMODARAN_MATURE_ERP_DECIMAL = 0.055`` fallback per Week 10
Sprint B narrowed scope).

Scope:
- ``fetch_annual_erp(year)`` — annual file, decimal Implied ERP.
- ``fetch_monthly_implied_erp(year, month)`` — monthly file row for
  start-of-month, Week 10 Sprint B Commit 2.
"""

from __future__ import annotations

import io
from calendar import month_abbr
from dataclasses import dataclass
from datetime import date
from typing import TYPE_CHECKING, cast

import httpx
import openpyxl
import structlog
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential_jitter,
)

from sonar.connectors.cache import ConnectorCache

if TYPE_CHECKING:
    from pathlib import Path

log = structlog.get_logger()

HISTIMPL_URL: str = "https://pages.stern.nyu.edu/~adamodar/pc/datasets/histimpl.xlsx"
# Refresh quarterly (Damodaran republishes annually but timestamp drifts).
HISTIMPL_TTL_SECONDS: int = 90 * 24 * 3600

# Column name (canonical per 2024 snapshot; parser reads by name not index).
COL_YEAR = "Year"
COL_IMPLIED_FCFE_SUSTAINABLE = "Implied Premium (FCFE with sustainable Payout)"
COL_IMPLIED_FCFE = "Implied ERP (FCFE)"
COL_SHEET = "Historical Impl Premiums"

# Monthly implied ERP archive (Week 10 Sprint B Commit 2).
IMPLPREM_URL_TEMPLATE: str = "https://pages.stern.nyu.edu/~adamodar/pc/implprem/ERP{mon}{yy}.xlsx"
# Refresh monthly — new file lands ~2mo after the row date it adds.
IMPLPREM_TTL_SECONDS: int = 30 * 24 * 3600
# Empirically found via Week 10 Sprint B probe 2026-04-22 — the
# monthly workbook's sheet holding the month-by-month implied ERP
# series from Sep 2008 onwards.
IMPLPREM_SHEET: str = "Implied ERP (Monthly from 9-08)"
# Monthly-file column headers (probed 2026-04-22 on ERPFeb26.xlsx).
COL_MONTH_START = "Start of month"
COL_MONTH_SP500 = "S&P 500"
COL_MONTH_TBOND_RATE = "T.Bond Rate"
COL_MONTH_ERP_SUSTAINABLE = "ERP (T12 m with sustainable payout)"
COL_MONTH_ERP_T12M = "ERP (T12m)"
# Max months the resolver will walk backward from the target to find
# a published monthly workbook. Damodaran's publication lag is ~2mo,
# so 6 gives generous headroom for unexpected delays.
IMPLPREM_LOOKBACK_MONTHS: int = 6


@dataclass(frozen=True, slots=True)
class DamodaranERPRow:
    """Annual implied ERP snapshot from Damodaran histimpl.xlsx."""

    year: int
    implied_erp_decimal: float
    source_column: str  # which column was used (we fall back on FCFE if sustainable missing)


@dataclass(frozen=True, slots=True)
class DamodaranMonthlyERPRow:
    """Monthly implied ERP snapshot from Damodaran implprem/ERPMMMYY.xlsx.

    ``implied_erp_decimal`` carries the preferred "ERP (T12 m with
    sustainable payout)" column (spec §4 primary convention); the plain
    "ERP (T12m)" column is retained under ``implied_erp_t12m_decimal``
    as a fallback + cross-check.
    """

    start_of_month: date
    implied_erp_decimal: float
    implied_erp_t12m_decimal: float | None
    sp500_level: float
    tbond_rate_decimal: float
    source_file: str  # e.g. "ERPFeb26.xlsx" — resolved at fetch time


class DamodaranConnector:
    """L0 connector for Damodaran histimpl.xlsx + monthly ERP archive."""

    BASE_URL = HISTIMPL_URL
    CACHE_NAMESPACE = "damodaran:histimpl"
    CACHE_NAMESPACE_IMPLPREM = "damodaran:implprem"

    def __init__(self, cache_dir: str | Path, timeout: float = 60.0) -> None:
        self.cache = ConnectorCache(cache_dir)
        self.client = httpx.AsyncClient(timeout=timeout, follow_redirects=True)

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential_jitter(initial=2, max=30),
        retry=retry_if_exception_type((httpx.HTTPError, httpx.TimeoutException)),
    )
    async def _download(self) -> bytes:
        r = await self.client.get(self.BASE_URL)
        r.raise_for_status()
        return r.content

    async def fetch_raw_xlsx(self) -> bytes:
        cached = self.cache.get(self.CACHE_NAMESPACE)
        if cached is not None:
            log.debug("damodaran.cache_hit")
            return cast("bytes", cached)
        body = await self._download()
        self.cache.set(self.CACHE_NAMESPACE, body, ttl=HISTIMPL_TTL_SECONDS)
        log.info("damodaran.fetched", bytes=len(body))
        return body

    async def fetch_annual_erp(self, year: int) -> DamodaranERPRow | None:
        """Return the annual implied ERP row for ``year``, or ``None``.

        Prefers the FCFE-with-sustainable-payout column (spec §4 primary
        convention); falls back to the plain FCFE column. Returns
        ``None`` when the year is absent from the file.
        """
        body = await self.fetch_raw_xlsx()
        return _parse_year(body, year)

    # ------------------------------------------------------------------
    # Monthly implied ERP archive — Week 10 Sprint B Commit 2
    # ------------------------------------------------------------------

    async def _download_implprem(self, url: str) -> bytes:
        """Single-attempt download of a monthly implied-ERP xlsx.

        Returns ``b""`` on 404 so the caller can walk backward through
        month candidates without bubbling a retry (Damodaran's 404s
        are a definite signal, not a transient error).
        """
        try:
            r = await self.client.get(url)
        except httpx.HTTPError as exc:
            log.warning("damodaran.implprem.http_error", url=url, error=str(exc))
            return b""
        if r.status_code == 404:
            return b""
        r.raise_for_status()
        return r.content

    async def fetch_monthly_implied_erp(
        self, year: int, month: int
    ) -> DamodaranMonthlyERPRow | None:
        """Return the start-of-month implied ERP row for ``(year, month)``.

        Resolution:

        1. Walk backward ≤ :data:`IMPLPREM_LOOKBACK_MONTHS` months from
           the target to locate the first published ``ERP{MonAbbr}{YY}.xlsx``
           file (cached 30d once resolved).
        2. Parse the ``Implied ERP (Monthly from 9-08)`` sheet for the
           row whose ``Start of month`` matches ``(year, month, 1)``.
        3. Return ``None`` when no monthly file can be resolved OR the
           target month is absent from the workbook (e.g. pre 2008-09).

        Each monthly workbook contains *all* historical months back to
        2008-09, so a single recent file satisfies any lookup inside
        its history — we cache that body under
        :data:`CACHE_NAMESPACE_IMPLPREM` keyed by resolved filename so
        repeat calls for different ``(year, month)`` don't re-download.
        """
        if month < 1 or month > 12:
            msg = f"month out of range: {month}"
            raise ValueError(msg)

        body_and_filename = await self._resolve_latest_implprem(year, month)
        if body_and_filename is None:
            return None
        body, source_file = body_and_filename

        return _parse_monthly(body, year, month, source_file)

    async def _resolve_latest_implprem(
        self, target_year: int, target_month: int
    ) -> tuple[bytes, str] | None:
        """Walk backward from ``(target_year, target_month)`` to find a published file.

        Returns ``(body, filename)`` on hit or ``None`` when no file
        resolves within :data:`IMPLPREM_LOOKBACK_MONTHS`.

        Caches the most recently resolved body under a single namespace
        key (:data:`CACHE_NAMESPACE_IMPLPREM`) so repeat lookups for
        different ``(year, month)`` share one download — each Damodaran
        monthly workbook carries the full month-by-month history back
        to 2008-09, so any recent file satisfies any historical query
        inside its range.
        """
        cached = self.cache.get(self.CACHE_NAMESPACE_IMPLPREM)
        if isinstance(cached, tuple) and len(cached) == 2:
            body_cached, filename_cached = cached
            if isinstance(body_cached, (bytes, bytearray)) and isinstance(filename_cached, str):
                log.debug("damodaran.implprem.cache_hit", file=filename_cached)
                return bytes(body_cached), filename_cached

        y, m = target_year, target_month
        for _ in range(IMPLPREM_LOOKBACK_MONTHS):
            filename = f"ERP{month_abbr[m]}{y % 100:02d}.xlsx"
            url = IMPLPREM_URL_TEMPLATE.format(mon=month_abbr[m], yy=f"{y % 100:02d}")
            body = await self._download_implprem(url)
            if body:
                self.cache.set(
                    self.CACHE_NAMESPACE_IMPLPREM,
                    (body, filename),
                    ttl=IMPLPREM_TTL_SECONDS,
                )
                log.info("damodaran.implprem.fetched", file=filename, bytes=len(body))
                return body, filename
            log.debug("damodaran.implprem.not_found", file=filename)
            # Step backward one month.
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        return None

    async def aclose(self) -> None:
        await self.client.aclose()
        self.cache.close()


def _parse_year(body: bytes, year: int) -> DamodaranERPRow | None:
    wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    ws = wb[COL_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    # Find the header row (contains "Year" in column 0).
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if row and row[0] == COL_YEAR:
            header_idx = i
            break
    if header_idx is None:
        return None
    header = list(rows[header_idx])
    try:
        col_sust = header.index(COL_IMPLIED_FCFE_SUSTAINABLE)
    except ValueError:
        col_sust = -1
    try:
        col_fcfe = header.index(COL_IMPLIED_FCFE)
    except ValueError:
        col_fcfe = -1

    for row in rows[header_idx + 1 :]:
        if not row or row[0] != year:
            continue
        # Prefer sustainable payout column; fall back to FCFE.
        if col_sust >= 0 and col_sust < len(row) and row[col_sust] is not None:
            return DamodaranERPRow(
                year=year,
                implied_erp_decimal=float(row[col_sust]),
                source_column=COL_IMPLIED_FCFE_SUSTAINABLE,
            )
        if col_fcfe >= 0 and col_fcfe < len(row) and row[col_fcfe] is not None:
            return DamodaranERPRow(
                year=year,
                implied_erp_decimal=float(row[col_fcfe]),
                source_column=COL_IMPLIED_FCFE,
            )
        return None
    return None


def _parse_monthly(
    body: bytes,
    target_year: int,
    target_month: int,
    source_file: str,
) -> DamodaranMonthlyERPRow | None:
    """Parse a monthly ERP workbook and return the row for ``(year, month)``.

    Returns ``None`` when the sheet / row / sustainable-payout column is
    absent (first months pre-2011 have sparse rows with no sustainable
    column populated — caller falls back to ``ERP (T12m)``).
    """
    wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    if IMPLPREM_SHEET not in wb.sheetnames:
        log.warning(
            "damodaran.implprem.sheet_missing",
            file=source_file,
            expected=IMPLPREM_SHEET,
            actual=wb.sheetnames,
        )
        return None
    ws = wb[IMPLPREM_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if row and row[0] == COL_MONTH_START:
            header_idx = i
            break
    if header_idx is None:
        return None
    header = list(rows[header_idx])

    def _col(name: str) -> int:
        try:
            return header.index(name)
        except ValueError:
            return -1

    col_start = _col(COL_MONTH_START)
    col_sp500 = _col(COL_MONTH_SP500)
    col_tbond = _col(COL_MONTH_TBOND_RATE)
    col_sust = _col(COL_MONTH_ERP_SUSTAINABLE)
    col_t12m = _col(COL_MONTH_ERP_T12M)

    for row in rows[header_idx + 1 :]:
        if not row or row[col_start] is None:
            continue
        cell = row[col_start]
        # The Start-of-month column is a datetime in openpyxl output.
        row_year: int | None = None
        row_month: int | None = None
        if hasattr(cell, "year") and hasattr(cell, "month"):
            row_year = int(cell.year)
            row_month = int(cell.month)
        if row_year != target_year or row_month != target_month:
            continue

        sust = _cell_to_float(row, col_sust)
        t12m = _cell_to_float(row, col_t12m)
        if sust is None and t12m is None:
            # No ERP data for this month — pre-2011 rows often have
            # only the SP500/Tbond columns.
            return None
        primary = sust if sust is not None else t12m
        assert primary is not None  # guarded above
        sp500 = _cell_to_float(row, col_sp500) or 0.0
        tbond = _cell_to_float(row, col_tbond) or 0.0
        return DamodaranMonthlyERPRow(
            start_of_month=date(target_year, target_month, 1),
            implied_erp_decimal=primary,
            implied_erp_t12m_decimal=t12m,
            sp500_level=sp500,
            tbond_rate_decimal=tbond,
            source_file=source_file,
        )
    return None


def _cell_to_float(row: tuple[object, ...], col: int) -> float | None:
    """Helper — read a row cell as float, returning ``None`` on miss.

    Kept module-level to avoid the per-row closure (Ruff B023) over the
    iterating ``row`` variable inside :func:`_parse_monthly`.
    """
    if col < 0 or col >= len(row):
        return None
    val = row[col]
    if val is None:
        return None
    try:
        return float(val)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
