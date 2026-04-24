"""Unit tests for :mod:`sonar.connectors.boe_yield_curves` (Sprint Q.2).

Tests build a synthetic ``glcinflationddata.zip`` in-memory that
mirrors the BoE workbook layout (sheet ``4. spot curve`` with a
``years:`` header row at row 4 and daily data rows starting at row
6), feed it through the connector's cache layer, and assert the
parsed :class:`BoeBeiSpotObservation` list is correct. No network
access.
"""

from __future__ import annotations

import asyncio
import io
import zipfile
from datetime import date
from typing import TYPE_CHECKING

import openpyxl
import pytest

from sonar.connectors.boe_yield_curves import (
    _INFLATION_ARCHIVE_FILE_BANDS,
    _NOMINAL_ARCHIVE_FILE_BANDS,
    BoeBeiSpotObservation,
    BoeNominalSpotObservation,
    BoeYieldCurvesConnector,
    _select_archive_files,
)
from sonar.overlays.exceptions import DataUnavailableError

if TYPE_CHECKING:
    from pathlib import Path


def _build_spot_curve_xlsx() -> bytes:
    """Return a minimal in-memory xlsx matching BoE's spot-curve layout.

    Layout: sheet ``4. spot curve`` with the tenor-header row 4 carrying
    {2.5, 3.0, ..., 10.0, ..., 30.0} — we just need 5Y (idx 6) and 10Y
    (idx 16) to be parseable.
    """
    wb = openpyxl.Workbook()
    # Default sheet gets renamed out of the way.
    ws = wb.active
    ws.title = "info"
    ws["A1"] = "info sheet"

    spot = wb.create_sheet("4. spot curve")
    # Row 1: title
    spot["A1"] = "UK implied inflation spot curve"
    # Row 2 empty.
    # Row 3: label
    spot["A3"] = "Maturity"
    # Row 4: tenor header — col A label, cols B.. = 2.5, 3.0, 3.5, ..., 10.0, ...
    spot["A4"] = "years:"
    tenors = [2.5 + 0.5 * i for i in range(60)]  # 2.5 .. 32.0
    for i, t in enumerate(tenors):
        spot.cell(row=4, column=2 + i, value=t)
    # Row 5: refresh
    spot["A5"] = "Refresh"
    # Row 6+: data. Seed 3 consecutive weekdays.
    data_rows = [
        (date(2024, 7, 1), 3.00, 3.30),  # 5Y=3.00%, 10Y=3.30%
        (date(2024, 7, 2), 3.05, 3.25),
        (date(2024, 7, 3), 2.95, 3.20),
    ]
    for row_i, (d, y5, y10) in enumerate(data_rows):
        row_idx = 6 + row_i
        spot.cell(row=row_idx, column=1, value=d)
        spot.cell(row=row_idx, column=2 + 5, value=y5)  # col 7 = B+5 = 5Y
        spot.cell(row=row_idx, column=2 + 15, value=y10)  # col 17 = 10Y
        # Fill 15Y/20Y/30Y too so _iter_data_rows emits the observation
        spot.cell(row=row_idx, column=2 + 25, value=3.15)
        spot.cell(row=row_idx, column=2 + 35, value=3.10)
        spot.cell(row=row_idx, column=2 + 55, value=3.05)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_archive_zip(xlsx_bytes: bytes) -> bytes:
    """Pack the synthetic xlsx into a zip under the 2024-band archive name."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("GLC Inflation daily data_2016 to 2024.xlsx", xlsx_bytes)
    return buf.getvalue()


@pytest.fixture
def connector(tmp_path: Path) -> BoeYieldCurvesConnector:
    return BoeYieldCurvesConnector(cache_dir=tmp_path / "boe_yc_cache")


def test_select_archive_files_single_band() -> None:
    names = _select_archive_files(date(2022, 3, 1), date(2023, 6, 1), _INFLATION_ARCHIVE_FILE_BANDS)
    assert names == ["GLC Inflation daily data_2016 to 2024.xlsx"]


def test_select_archive_files_spans_present_band() -> None:
    names = _select_archive_files(
        date(2024, 12, 1), date(2026, 2, 1), _INFLATION_ARCHIVE_FILE_BANDS
    )
    assert "GLC Inflation daily data_2016 to 2024.xlsx" in names
    assert "GLC Inflation daily data_2025 to present.xlsx" in names


def test_select_archive_files_skips_out_of_band() -> None:
    names = _select_archive_files(
        date(2022, 1, 1), date(2022, 12, 31), _INFLATION_ARCHIVE_FILE_BANDS
    )
    # Only the 2016-2024 file should match; historical 1985..2015 + 2025-present skipped.
    assert names == ["GLC Inflation daily data_2016 to 2024.xlsx"]


def test_fetch_inflation_spot_curve_parses_cached_archive(
    connector: BoeYieldCurvesConnector,
) -> None:
    """End-to-end: seed cache with synthetic zip → connector returns observations."""
    archive = _build_archive_zip(_build_spot_curve_xlsx())
    connector.cache.set(
        f"{connector.CACHE_NAMESPACE}:archive:glcinflationddata.zip",
        archive,
        ttl=86400,
    )
    try:
        obs = asyncio.run(connector.fetch_inflation_spot_curve(date(2024, 7, 1), date(2024, 7, 3)))
    finally:
        asyncio.run(connector.aclose())
    assert len(obs) == 3
    assert all(isinstance(o, BoeBeiSpotObservation) for o in obs)
    assert all(o.country_code == "GB" for o in obs)
    assert all(o.source == "BOE_GLC_INFLATION" for o in obs)
    # First row: 5Y=3.00% → 0.03; 10Y=3.30% → 0.033.
    first = obs[0]
    assert first.observation_date == date(2024, 7, 1)
    assert first.tenors["5Y"] == pytest.approx(0.0300)
    assert first.tenors["10Y"] == pytest.approx(0.0330)
    # Full tenor list covered (5Y..30Y).
    assert set(first.tenors) == {"5Y", "10Y", "15Y", "20Y", "30Y"}


def test_fetch_inflation_spot_curve_filters_by_date_window(
    connector: BoeYieldCurvesConnector,
) -> None:
    """Rows outside the requested date window are excluded."""
    archive = _build_archive_zip(_build_spot_curve_xlsx())
    connector.cache.set(
        f"{connector.CACHE_NAMESPACE}:archive:glcinflationddata.zip",
        archive,
        ttl=86400,
    )
    try:
        obs = asyncio.run(connector.fetch_inflation_spot_curve(date(2024, 7, 2), date(2024, 7, 2)))
    finally:
        asyncio.run(connector.aclose())
    assert [o.observation_date for o in obs] == [date(2024, 7, 2)]


def test_fetch_inflation_spot_curve_raises_when_window_has_no_rows(
    connector: BoeYieldCurvesConnector,
) -> None:
    """Empty-window parse → DataUnavailableError, not empty list."""
    archive = _build_archive_zip(_build_spot_curve_xlsx())
    connector.cache.set(
        f"{connector.CACHE_NAMESPACE}:archive:glcinflationddata.zip",
        archive,
        ttl=86400,
    )
    try:
        with pytest.raises(DataUnavailableError):
            asyncio.run(connector.fetch_inflation_spot_curve(date(2023, 1, 1), date(2023, 1, 31)))
    finally:
        asyncio.run(connector.aclose())


def test_fetch_inflation_spot_curve_rejects_inverted_window(
    connector: BoeYieldCurvesConnector,
) -> None:
    try:
        with pytest.raises(ValueError, match="date_start"):
            asyncio.run(connector.fetch_inflation_spot_curve(date(2024, 7, 3), date(2024, 7, 1)))
    finally:
        asyncio.run(connector.aclose())


# -----------------------------------------------------------------------------
# Sprint P.2 — nominal spot curve coverage.
# -----------------------------------------------------------------------------


def _build_nominal_spot_curve_xlsx() -> bytes:
    """Synthetic xlsx mirroring the BoE **nominal** workbook layout.

    Unlike the inflation workbook, the nominal header row starts at
    0.5Y in column B, so columns step at ``col = 2 * years`` for whole-
    year tenors (2Y→4, 3Y→6, 5Y→10, 7Y→14, 10Y→20, 15Y→30, 20Y→40,
    30Y→60). We seed the 8 tenors in
    :data:`NOMINAL_SPOT_CURVE_TENOR_COLUMNS` so the connector's default
    ``_DEFAULT_NOMINAL_TENORS`` set parses cleanly.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "info"
    ws["A1"] = "info sheet"

    spot = wb.create_sheet("4. spot curve")
    spot["A1"] = "UK nominal spot curve"
    spot["A3"] = "Maturity"
    spot["A4"] = "years:"
    tenors_header = [0.5 + 0.5 * i for i in range(80)]  # 0.5 .. 40.0
    for i, t in enumerate(tenors_header):
        spot.cell(row=4, column=2 + i, value=t)
    spot["A5"] = "Refresh"

    # Seed 2 consecutive weekdays with an upward-sloping curve across the
    # 8 default tenors (values in percent).
    default_tenor_years = (2.0, 3.0, 5.0, 7.0, 10.0, 15.0, 20.0, 30.0)
    data_rows = [
        (date(2024, 7, 1), [4.10, 4.20, 4.40, 4.55, 4.65, 4.75, 4.80, 4.85]),
        (date(2024, 7, 2), [4.15, 4.22, 4.43, 4.58, 4.68, 4.78, 4.82, 4.88]),
    ]
    for row_i, (d, yields_pct) in enumerate(data_rows):
        row_idx = 6 + row_i
        spot.cell(row=row_idx, column=1, value=d)
        for tenor_y, pct in zip(default_tenor_years, yields_pct, strict=True):
            spot.cell(row=row_idx, column=1 + int(2 * tenor_y), value=pct)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_nominal_archive_zip(xlsx_bytes: bytes) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("GLC Nominal daily data_2016 to 2024.xlsx", xlsx_bytes)
    return buf.getvalue()


def test_select_nominal_archive_files_covers_2020_to_present() -> None:
    names = _select_archive_files(date(2020, 1, 1), date(2026, 4, 24), _NOMINAL_ARCHIVE_FILE_BANDS)
    assert "GLC Nominal daily data_2016 to 2024.xlsx" in names
    assert "GLC Nominal daily data_2025 to present.xlsx" in names
    # Earlier bands (1979-2015) excluded from the 2020-onwards request.
    assert "GLC Nominal daily data_2005 to 2015.xlsx" not in names


def test_fetch_nominal_spot_curve_parses_cached_archive(
    connector: BoeYieldCurvesConnector,
) -> None:
    """End-to-end: seed cache with synthetic nominal zip → observations."""
    archive = _build_nominal_archive_zip(_build_nominal_spot_curve_xlsx())
    connector.cache.set(
        f"{connector.CACHE_NAMESPACE}:archive:glcnominalddata.zip",
        archive,
        ttl=86400,
    )
    try:
        obs = asyncio.run(connector.fetch_nominal_spot_curve(date(2024, 7, 1), date(2024, 7, 2)))
    finally:
        asyncio.run(connector.aclose())
    assert len(obs) == 2
    assert all(isinstance(o, BoeNominalSpotObservation) for o in obs)
    assert all(o.country_code == "GB" for o in obs)
    assert all(o.source == "BOE_GLC_NOMINAL" for o in obs)
    first = obs[0]
    assert first.observation_date == date(2024, 7, 1)
    # 5Y = 4.40 % → 0.044; 10Y = 4.65 % → 0.0465.
    assert first.tenors["5Y"] == pytest.approx(0.0440)
    assert first.tenors["10Y"] == pytest.approx(0.0465)
    # Full 8-tenor default set parsed.
    assert set(first.tenors) == {"2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "30Y"}


def test_fetch_nominal_spot_curve_rejects_inverted_window(
    connector: BoeYieldCurvesConnector,
) -> None:
    try:
        with pytest.raises(ValueError, match="date_start"):
            asyncio.run(connector.fetch_nominal_spot_curve(date(2024, 7, 3), date(2024, 7, 1)))
    finally:
        asyncio.run(connector.aclose())
